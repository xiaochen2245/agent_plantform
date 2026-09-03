"""
工程量清单与电子表格 (XLSX / XLS / XLSM) 解析器
基于 openpyxl 实现多工作表遍历、合并单元格前向传播对齐 (Forward-Filling)、
多级复合表头结构识别与高保真 Markdown / Proposition 双重视角重构。
内置纯 Python 原生 XML 归档解压降级引擎，实现零外部依赖弹性运行。
"""

import io
import re
import uuid
import xml.etree.ElementTree as ET
import zipfile
from typing import Any, Dict, List, Optional, Set, Tuple

try:
    import openpyxl
    from openpyxl.worksheet.worksheet import Worksheet
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from app.schemas.ast import (
    ASTBlockType,
    ASTNode,
    DocumentSourceType,
    TableCell,
    TableData,
    UnifiedDocumentAST,
)
from .base import BaseParser, EmptyDocumentError, MalformedDocumentError


class XLSXParser(BaseParser):
    """
    XLSX 造价清单与复合电子表格解析适配器。
    精准还原跨行跨列合并单元格，解决传统表格解析器在单元格合并处产生数据空洞的痛点。
    """

    HEADER_KEYWORDS = {
        "序号", "编号", "项目编码", "项目名称", "项目特征", "项目特征描述",
        "计量单位", "工程量", "工程数量", "金额", "单价", "综合单价", "合价",
        "暂估价", "暂列金额", "规费", "税金", "措施项目", "备注"
    }

    @property
    def supported_extensions(self) -> List[str]:
        return [".xlsx", ".xlsm", ".xltx", ".xltm"]

    @property
    def source_type(self) -> DocumentSourceType:
        return DocumentSourceType.XLSX

    async def parse(
        self,
        content: bytes,
        file_name: str,
        tenant_id: str = "default",
        document_id: Optional[str] = None,
        **kwargs: Any,
    ) -> UnifiedDocumentAST:
        """异步解析 XLSX 字节流"""
        doc_id = document_id or f"doc_{uuid.uuid4().hex[:12]}"

        if not content:
            raise EmptyDocumentError(f"XLSX 文件 '{file_name}' 字节流为空", file_name=file_name)

        if HAS_OPENPYXL:
            try:
                # 使用 data_only=True 获取公式计算后的最终值，而非公式文本
                wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=False)
            except Exception as e:
                # 尝试纯 XML 降级
                return self._parse_pure_xml_fallback(content, file_name, tenant_id, doc_id)

            sheet_names = wb.sheetnames
            if not sheet_names:
                raise EmptyDocumentError(f"XLSX 工作簿 '{file_name}' 无任何有效工作表 (Sheets)", file_name=file_name)

            nodes: List[ASTNode] = []
            total_sheets = len(sheet_names)

            for sheet_idx, sname in enumerate(sheet_names, start=1):
                sheet: Worksheet = wb[sname]
                sheet_nodes = self._parse_single_sheet(sheet, sname, sheet_idx)
                nodes.extend(sheet_nodes)

            wb.close()

            if not nodes:
                nodes.append(
                    ASTNode(
                        block_id=self.generate_block_id("xlsx_empty"),
                        block_type=ASTBlockType.PARAGRAPH,
                        text_content="[工作簿包含的工作表均为空白或无有效数据]",
                        page_or_sheet=sheet_names[0] if sheet_names else "Sheet1"
                    )
                )

            return UnifiedDocumentAST(
                document_id=doc_id,
                tenant_id=tenant_id,
                file_name=file_name,
                source_type=self.source_type,
                total_pages_or_sheets=total_sheets,
                nodes=nodes,
                metadata={
                    "parser": "XLSXParser",
                    "sheet_names": sheet_names,
                    "total_sheets": total_sheets,
                    "extracted_nodes_count": len(nodes),
                }
            )
        else:
            return self._parse_pure_xml_fallback(content, file_name, tenant_id, doc_id)

    def _parse_single_sheet(
        self,
        sheet: Any,
        sheet_name: str,
        sheet_idx: int
    ) -> List[ASTNode]:
        """解析单个工作表，区分大标题、多级表头与合并表格体"""
        nodes: List[ASTNode] = []
        max_row = sheet.max_row or 0
        max_col = sheet.max_column or 0

        if max_row == 0 or max_col == 0:
            return nodes

        # 1. 扫描所有合并单元格范围并建立映射
        # (r, c) -> (val, row_span, col_span, is_top_left)
        merged_map: Dict[Tuple[int, int], Tuple[Any, int, int, bool]] = {}
        top_left_cells: Set[Tuple[int, int]] = set()

        for rng in list(sheet.merged_cells.ranges):
            min_r, min_c, max_r, max_c = rng.min_row, rng.min_col, rng.max_row, rng.max_col
            tl_val = sheet.cell(row=min_r, column=min_c).value
            row_span = max_r - min_r + 1
            col_span = max_c - min_c + 1
            top_left_cells.add((min_r, min_c))

            for r in range(min_r, max_r + 1):
                for c in range(min_c, max_c + 1):
                    is_tl = (r == min_r and c == min_c)
                    merged_map[(r, c)] = (tl_val, row_span, col_span, is_tl)

        # 2. 构建前向填充后的全量密集网格 (Dense Grid) 与 TableCell 集合
        grid_data: List[List[str]] = []
        table_cells: List[TableCell] = []

        for r in range(1, max_row + 1):
            row_vals: List[str] = []
            for c in range(1, max_col + 1):
                if (r, c) in merged_map:
                    val, r_span, c_span, is_tl = merged_map[(r, c)]
                    s_val = self._format_cell_value(val)
                    row_vals.append(s_val)
                    if is_tl:
                        table_cells.append(
                            TableCell(
                                row=r - 1,
                                col=c - 1,
                                row_span=r_span,
                                col_span=c_span,
                                text=s_val,
                                is_header=False
                            )
                        )
                else:
                    raw_val = sheet.cell(row=r, column=c).value
                    s_val = self._format_cell_value(raw_val)
                    row_vals.append(s_val)
                    if s_val:
                        table_cells.append(
                            TableCell(
                                row=r - 1,
                                col=c - 1,
                                row_span=1,
                                col_span=1,
                                text=s_val,
                                is_header=False
                            )
                        )
            grid_data.append(row_vals)

        # 过滤全空行
        non_empty_rows: List[Tuple[int, List[str]]] = [
            (idx, row) for idx, row in enumerate(grid_data) if any(cell.strip() for cell in row)
        ]
        if not non_empty_rows:
            return nodes

        # 3. 识别顶层标题、表头行与数据体行
        start_row_idx = 0
        sheet_title_block: Optional[ASTNode] = None

        first_row_idx, first_row_vals = non_empty_rows[0]
        non_empty_cells_count = sum(1 for c in first_row_vals if c.strip())
        if non_empty_cells_count <= 2 and len("".join(first_row_vals).strip()) > 3:
            title_text = "".join(first_row_vals).strip()
            sheet_title_block = ASTNode(
                block_id=self.generate_block_id("xlsx_head"),
                block_type=ASTBlockType.HEADING,
                level=2,
                section_path=[sheet_name, title_text],
                text_content=title_text,
                page_or_sheet=sheet_name,
                extra_metadata={"is_sheet_title": True}
            )
            nodes.append(sheet_title_block)
            start_row_idx = 1

        # 4. 识别表头行 (Header Rows)
        header_rows: List[List[str]] = []
        body_rows: List[List[str]] = []
        in_header_phase = True

        for i in range(start_row_idx, len(non_empty_rows)):
            _, row = non_empty_rows[i]
            row_text_set = {c.strip() for c in row if c.strip()}
            intersection = row_text_set.intersection(self.HEADER_KEYWORDS)
            
            if in_header_phase and (intersection or len(header_rows) < 2):
                header_rows.append(row)
                if len(header_rows) >= 3 and not intersection:
                    in_header_phase = False
            else:
                in_header_phase = False
                body_rows.append(row)

        if not header_rows and body_rows:
            header_rows.append(body_rows.pop(0))

        # 5. 生成 Markdown 表格与行命题
        section_path = [sheet_name]
        if sheet_title_block:
            section_path.append(sheet_title_block.text_content)

        markdown_table = self.build_table_markdown(headers=header_rows, rows=body_rows)
        table_summary = f"工作表: {sheet_name}, 包含 {len(header_rows)} 级表头与 {len(body_rows)} 行造价/清单记录。"

        table_node = ASTNode(
            block_id=self.generate_block_id("xlsx_tbl"),
            block_type=ASTBlockType.TABLE,
            section_path=section_path,
            text_content=markdown_table,
            table_data=TableData(
                headers=header_rows,
                rows=body_rows,
                cells=table_cells,
                markdown=markdown_table,
                summary=table_summary
            ),
            page_or_sheet=sheet_name,
            extra_metadata={
                "sheet_name": sheet_name,
                "sheet_index": sheet_idx,
                "total_rows": len(grid_data),
                "total_cols": max_col,
            }
        )
        nodes.append(table_node)

        return nodes

    def _parse_pure_xml_fallback(
        self,
        content: bytes,
        file_name: str,
        tenant_id: str,
        doc_id: str
    ) -> UnifiedDocumentAST:
        """纯 Python 遍历 xl/worksheets/ 归档的零外部依赖降级引擎"""
        nodes: List[ASTNode] = []
        sheet_names: List[str] = []

        try:
            with zipfile.ZipFile(io.BytesIO(content), "r") as z:
                # 1. 读取共享字符串表
                shared_strings: List[str] = []
                if "xl/sharedStrings.xml" in z.namelist():
                    s_xml = z.read("xl/sharedStrings.xml")
                    s_root = ET.fromstring(s_xml)
                    for si in s_root.iter():
                        if si.tag.endswith("si"):
                            t_texts = [e.text for e in si.iter() if e.tag.endswith("t") and e.text]
                            shared_strings.append("".join(t_texts))

                # 2. 读取工作簿结构以映射真实 Sheet 别名
                sheet_name_map: Dict[str, str] = {}
                if "xl/workbook.xml" in z.namelist():
                    wb_xml = z.read("xl/workbook.xml")
                    wb_root = ET.fromstring(wb_xml)
                    for s_elem in wb_root.iter():
                        if s_elem.tag.endswith("sheet"):
                            s_name = s_elem.get("name")
                            s_id = s_elem.get("sheetId")
                            if s_name and s_id:
                                sheet_name_map[f"sheet{s_id}.xml"] = s_name
                                sheet_names.append(s_name)

                # 辅助函数: 将 Excel 单元格坐标 (如 'C12') 解析为 (row, col) 1-based
                def parse_cell_ref(ref: str) -> Tuple[int, int]:
                    m = re.match(r"([A-Za-z]+)(\d+)", ref)
                    if not m:
                        return 1, 1
                    col_letters, row_str = m.group(1).upper(), m.group(2)
                    col_num = 0
                    for ch in col_letters:
                        col_num = col_num * 26 + (ord(ch) - ord("A") + 1)
                    return int(row_str), col_num

                # 3. 读取各工作表
                sheet_files = sorted(
                    [n for n in z.namelist() if re.match(r"^xl/worksheets/sheet\d+\.xml$", n)],
                    key=lambda x: int(re.search(r"\d+", x).group(0))
                )

                for sf in sheet_files:
                    xml_basename = sf.split("/")[-1]
                    sname = sheet_name_map.get(xml_basename, xml_basename.replace(".xml", "").capitalize())
                    ws_xml = z.read(sf)
                    ws_root = ET.fromstring(ws_xml)

                    # 解析合并单元格范围
                    merged_spans: List[Tuple[int, int, int, int]] = []
                    for elem in ws_root.iter():
                        if elem.tag.endswith("mergeCell"):
                            ref = elem.get("ref", "")
                            if ":" in ref:
                                start_ref, end_ref = ref.split(":")
                                r1, c1 = parse_cell_ref(start_ref)
                                r2, c2 = parse_cell_ref(end_ref)
                                merged_spans.append((r1, c1, r2, c2))

                    # 提取单元格值
                    raw_cells: Dict[Tuple[int, int], str] = {}
                    max_r, max_c = 0, 0

                    for r_elem in ws_root.iter():
                        if not r_elem.tag.endswith("row"):
                            continue
                        r_attr = r_elem.get("r")
                        r_num = int(r_attr) if r_attr and r_attr.isdigit() else 0

                        for c_elem in r_elem:
                            if not c_elem.tag.endswith("c"):
                                continue
                            c_ref = c_elem.get("r", "")
                            if c_ref:
                                r_idx, c_idx = parse_cell_ref(c_ref)
                            else:
                                r_idx = r_num
                                c_idx = max_c + 1

                            max_r = max(max_r, r_idx)
                            max_c = max(max_c, c_idx)

                            t_attr = c_elem.get("t")
                            val_str = ""

                            if t_attr == "s":
                                v_texts = [e.text for e in c_elem.iter() if e.tag.endswith("v") and e.text]
                                v_raw = v_texts[0] if v_texts else ""
                                if v_raw.isdigit():
                                    s_idx_val = int(v_raw)
                                    val_str = shared_strings[s_idx_val] if s_idx_val < len(shared_strings) else v_raw
                                else:
                                    val_str = v_raw
                            elif t_attr == "inlineStr":
                                t_texts = [e.text for e in c_elem.iter() if e.tag.endswith("t") and e.text]
                                val_str = "".join(t_texts)
                            else:
                                v_texts = [e.text for e in c_elem.iter() if (e.tag.endswith("v") or e.tag.endswith("t")) and e.text]
                                val_str = v_texts[0] if v_texts else ""

                            raw_cells[(r_idx, c_idx)] = val_str.strip()

                    if not raw_cells:
                        continue

                    # 前向填充合并单元格 (Forward-Filling)
                    for r1, c1, r2, c2 in merged_spans:
                        top_left_val = raw_cells.get((r1, c1), "")
                        for r in range(r1, r2 + 1):
                            for c in range(c1, c2 + 1):
                                raw_cells[(r, c)] = top_left_val

                    # 组装 2D 矩阵
                    grid: List[List[str]] = []
                    for r in range(1, max_r + 1):
                        row_vals: List[str] = []
                        for c in range(1, max_c + 1):
                            v = raw_cells.get((r, c), "")
                            row_vals.append(v)
                        if any(cell.strip() for cell in row_vals):
                            grid.append(row_vals)

                    if not grid:
                        continue

                    # 检查首行是否为大标题 (Banner Title)
                    first_row = grid[0]
                    non_empty_cells = [c.strip() for c in first_row if c.strip()]
                    distinct_texts = set(non_empty_cells)
                    start_idx = 0
                    sheet_title_block = None

                    is_banner = (
                        (len(distinct_texts) == 1 and len(list(distinct_texts)[0]) > 3)
                        or (len(non_empty_cells) <= 2 and len("".join(first_row).strip()) > 3)
                    )

                    if is_banner:
                        title_text = list(distinct_texts)[0] if distinct_texts else "".join(first_row).strip()
                        sheet_title_block = ASTNode(
                            block_id=self.generate_block_id("xlsx_head_fb"),
                            block_type=ASTBlockType.HEADING,
                            level=2,
                            section_path=[sname, title_text],
                            text_content=title_text,
                            page_or_sheet=sname,
                            extra_metadata={"is_sheet_title": True}
                        )
                        nodes.append(sheet_title_block)
                        start_idx = 1

                    table_rows = grid[start_idx:]
                    if table_rows:
                        headers = [table_rows[0]]
                        body = table_rows[1:] if len(table_rows) > 1 else []
                        md = self.build_table_markdown(headers=headers, rows=body)
                        sec_path = [sname]
                        if sheet_title_block:
                            sec_path.append(sheet_title_block.text_content)

                        table_cells = [
                            TableCell(row=r_i, col=c_i, row_span=1, col_span=1, text=c_val, is_header=(r_i == 0))
                            for r_i, r in enumerate(table_rows) for c_i, c_val in enumerate(r)
                        ]

                        nodes.append(
                            ASTNode(
                                block_id=self.generate_block_id("xlsx_tbl_fb"),
                                block_type=ASTBlockType.TABLE,
                                section_path=sec_path,
                                text_content=md,
                                table_data=TableData(
                                    headers=headers,
                                    rows=body,
                                    cells=table_cells,
                                    markdown=md,
                                    summary=f"工作表 {sname}，共 {len(body)} 行数据。"
                                ),
                                page_or_sheet=sname,
                                extra_metadata={"sheet_name": sname}
                            )
                        )
        except zipfile.BadZipFile as e:
            raise MalformedDocumentError(f"读取 XLSX XML 结构失败: {e}", file_name=file_name) from e
        except Exception as e:
            if isinstance(e, (EmptyDocumentError, MalformedDocumentError)):
                raise
            raise MalformedDocumentError(f"读取 XLSX XML 结构失败: {e}", file_name=file_name) from e

        if not nodes:
            nodes.append(
                ASTNode(
                    block_id=self.generate_block_id("xlsx_fb_empty"),
                    block_type=ASTBlockType.PARAGRAPH,
                    text_content="[工作簿中未检索到有效数据行]",
                    page_or_sheet="Sheet1"
                )
            )

        return UnifiedDocumentAST(
            document_id=doc_id,
            tenant_id=tenant_id,
            file_name=file_name,
            source_type=self.source_type,
            total_pages_or_sheets=max(len(sheet_names), 1),
            nodes=nodes,
            metadata={"parser": "XLSXParser", "is_pure_xml_fallback": True, "sheet_names": sheet_names}
        )

    def _format_cell_value(self, val: Any) -> str:
        """单元格数据格式化，处理浮点数与金额精度"""
        if val is None:
            return ""
        if isinstance(val, float):
            if val.is_integer():
                return str(int(val))
            return f"{val:.4f}".rstrip("0").rstrip(".")
        return str(val).strip()
